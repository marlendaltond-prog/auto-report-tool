#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
依赖管理模块 - 处理第三方库导入和可用性检查
"""

import logging
import os
from pathlib import Path
from typing import Dict, List, Any, Optional

# 创建依赖管理日志记录器
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 依赖库状态字典
DEPENDENCIES = {
    'pandas': False,
    'numpy': False,
    'openpyxl': False,
    'sqlalchemy': False,
    'jinja2': False,
    'reportlab': False,
    'requests': False,
    'schedule': False,
    'email': False
}

# 导入的模块引用
pd = None
np = None
openpyxl = None
sa = None
jinja2 = None
requests = None
schedule = None
smtplib = None
MIMEMultipart = None
MIMEText = None
MIMEBase = None
encoders = None
Font = None
Alignment = None
PatternFill = None
Border = None
Side = None
get_column_letter = None
BarChart = None
LineChart = None
PieChart = None
ScatterChart = None
Reference = None
Series = None
colors = None
letter = None
A4 = None
SimpleDocTemplate = None
Table = None
TableStyle = None
Paragraph = None
getSampleStyleSheet = None

def check_dependencies():
    """检查并导入所有第三方依赖库"""
    
    # 导入 pandas
    try:
        global pd
        import pandas as pd
        DEPENDENCIES['pandas'] = True
        logger.info("✅ 成功导入 pandas")
    except ImportError as e:
        logger.warning(f"❌ 导入 pandas 失败: {e}")
    
    # 导入 numpy
    try:
        global np
        import numpy as np
        DEPENDENCIES['numpy'] = True
        logger.info("✅ 成功导入 numpy")
    except ImportError as e:
        logger.warning(f"❌ 导入 numpy 失败: {e}")
    
    # 导入 openpyxl
    try:
        global openpyxl, Font, Alignment, PatternFill, Border, Side
        global get_column_letter, BarChart, LineChart, PieChart, ScatterChart, Reference, Series
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
        DEPENDENCIES['openpyxl'] = True
        logger.info("✅ 成功导入 openpyxl")
    except ImportError as e:
        logger.warning(f"❌ 导入 openpyxl 失败: {e}")
    
    # 导入 sqlalchemy
    try:
        global sa
        import sqlalchemy as sa
        DEPENDENCIES['sqlalchemy'] = True
        logger.info("✅ 成功导入 sqlalchemy")
    except ImportError as e:
        logger.warning(f"❌ 导入 sqlalchemy 失败: {e}")
    
    # 导入 jinja2
    try:
        global jinja2
        import jinja2
        DEPENDENCIES['jinja2'] = True
        logger.info("✅ 成功导入 jinja2")
    except ImportError as e:
        logger.warning(f"❌ 导入 jinja2 失败: {e}")
    
    # 导入 reportlab
    try:
        global colors, letter, A4, SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        DEPENDENCIES['reportlab'] = True
        logger.info("✅ 成功导入 reportlab")
    except ImportError as e:
        logger.warning(f"❌ 导入 reportlab 失败: {e}")
    
    # 导入 requests
    try:
        global requests
        import requests
        DEPENDENCIES['requests'] = True
        logger.info("✅ 成功导入 requests")
    except ImportError as e:
        logger.warning(f"❌ 导入 requests 失败: {e}")
    
    # 导入 schedule
    try:
        global schedule
        import schedule
        DEPENDENCIES['schedule'] = True
        logger.info("✅ 成功导入 schedule")
    except ImportError as e:
        logger.warning(f"❌ 导入 schedule 失败: {e}")
    
    # 导入 email模块
    try:
        global smtplib, MIMEMultipart, MIMEText, MIMEBase, encoders
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        DEPENDENCIES['email'] = True
        logger.info("✅ 成功导入 email 模块")
    except ImportError as e:
        logger.warning(f"❌ 导入 email 模块失败: {e}")
    
    # 验证核心库可用性
    core_libraries = ['pandas', 'numpy', 'openpyxl']
    available_core = [dep for dep in core_libraries if DEPENDENCIES[dep]]
    
    if len(available_core) == 0:
        logger.error("🚫 所有核心库都不可用，程序将无法正常运行")
        logger.error("请安装必要的依赖包: pip install pandas openpyxl numpy")
        raise ImportError("缺少核心依赖库，程序无法运行")
    elif len(available_core) < len(core_libraries):
        logger.warning(f"⚠️  部分核心库不可用（{len(available_core)}/{len(core_libraries)} 核心库可用）")
        missing_core = [dep for dep in core_libraries if not DEPENDENCIES[dep]]
        logger.warning(f"缺少的核心库: {', '.join(missing_core)}")
        logger.warning("建议安装完整依赖: pip install pandas openpyxl sqlalchemy jinja2 reportlab requests schedule")
    else:
        logger.info("🎉 所有核心库可用")
    
    # 记录所有依赖状态
    available_deps = [k for k, v in DEPENDENCIES.items() if v]
    missing_deps = [k for k, v in DEPENDENCIES.items() if not v]
    
    logger.info(f"📊 可用依赖 ({len(available_deps)}/{len(DEPENDENCIES)}): {', '.join(available_deps)}")
    if missing_deps:
        logger.info(f"❌ 缺失依赖 ({len(missing_deps)}): {', '.join(missing_deps)}")
    
    return DEPENDENCIES

def check_feature(feature_name: str, required_deps: List[str]) -> bool:
    """检查功能是否可用"""
    missing_deps = [dep for dep in required_deps if not DEPENDENCIES.get(dep, False)]
    if missing_deps:
        logger.warning(f"功能 '{feature_name}' 不可用，缺少依赖: {', '.join(missing_deps)}")
        return False
    return True

def require_feature(feature_name: str, required_deps: List[str]):
    """功能依赖检查装饰器"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            missing_deps = [dep for dep in required_deps if not DEPENDENCIES.get(dep, False)]
            if missing_deps:
                raise ImportError(
                    f"功能 '{feature_name}' 不可用，缺少依赖: {', '.join(missing_deps)}"
                )
            return func(*args, **kwargs)
        return wrapper
    return decorator

def get_missing_dependencies() -> List[str]:
    """获取缺失的依赖列表"""
    return [dep for dep, available in DEPENDENCIES.items() if not available]

def is_core_complete() -> bool:
    """检查核心依赖是否完整"""
    core_deps = ['pandas', 'numpy', 'openpyxl']
    return all(DEPENDENCIES.get(dep, False) for dep in core_deps)

# 执行依赖检查
if __name__ == "__main__":
    # 如果直接运行此文件，执行依赖检查
    print("🔍 检查依赖库...")
    check_dependencies()
    print("✅ 依赖检查完成")