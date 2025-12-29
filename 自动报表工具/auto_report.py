"""
自动化报表工具 - AutoReport Pro
支持多种数据源：Excel, CSV, SQL数据库, API
支持多种输出：Excel, PDF, HTML, Email, 云存储
"""
from datetime import datetime, timedelta
import os
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import logging
from dataclasses import dataclass, field
from abc import ABC, abstractmethod
import warnings
import base64
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
warnings.filterwarnings('ignore')

# 第三方库导入（需要安装）
# 核心库标记
pandas_available = True
numpy_available = True
openpyxl_available = True
sqlalchemy_available = True
jinja2_available = True
reportlab_available = True
requests_available = True
schedule_available = True
email_available = True

# 模块引用
pd = None
np = None
openpyxl = None
Font = None
Alignment = None
PatternFill = None
Border = None
Side = None
get_column_letter = None
BarChart = None
LineChart = None
PieChart = None
ScatterChart = None
Reference = None
Series = None
sa = None
smtplib = None
MIMEMultipart = None
MIMEText = None
MIMEBase = None
encoders = None
jinja2 = None
colors = None
letter = None
A4 = None
SimpleDocTemplate = None
Table = None
TableStyle = None
Paragraph = None
getSampleStyleSheet = None
json = None
requests = None
schedule = None

try:
    import pandas as pd
except ImportError as e:
    pandas_available = False
    logger.error(f"导入 pandas 失败: {e}")
    print(f"警告: 缺少依赖包 pandas: {e}")

try:
    import numpy as np
except ImportError as e:
    numpy_available = False
    logger.error(f"导入 numpy 失败: {e}")
    print(f"警告: 缺少依赖包 numpy: {e}")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
except ImportError as e:
    openpyxl_available = False
    logger.error(f"导入 openpyxl 失败: {e}")
    print(f"警告: 缺少依赖包 openpyxl: {e}")

try:
    import sqlalchemy as sa
except ImportError as e:
    sqlalchemy_available = False
    logger.error(f"导入 sqlalchemy 失败: {e}")
    print(f"警告: 缺少依赖包 sqlalchemy: {e}")

try:
    import jinja2
except ImportError as e:
    jinja2_available = False
    logger.error(f"导入 jinja2 失败: {e}")
    print(f"警告: 缺少依赖包 jinja2: {e}")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
except ImportError as e:
    reportlab_available = False
    logger.error(f"导入 reportlab 失败: {e}")
    print(f"警告: 缺少依赖包 reportlab: {e}")

try:
    import requests
except ImportError as e:
    requests_available = False
    logger.error(f"导入 requests 失败: {e}")
    print(f"警告: 缺少依赖包 requests: {e}")

try:
    import schedule
    import time
except ImportError as e:
    schedule_available = False
    logger.error(f"导入 schedule 失败: {e}")
    print(f"警告: 缺少依赖包 schedule: {e}")

try:
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
except ImportError as e:
    email_available = False
    logger.error(f"导入 email 模块失败: {e}")
    print(f"警告: 导入 email 模块失败: {e}")

try:
    import json
except ImportError as e:
    logger.error(f"导入 json 失败: {e}")
    print(f"警告: 导入 json 失败: {e}")

# 检查核心功能是否可用
core_libraries = [pandas_available, numpy_available, openpyxl_available]
third_party_available = all(core_libraries) or any(core_libraries)

if not third_party_available:
    logger.error("所有核心库都不可用，程序将无法正常运行")
    print("错误: 所有核心库都不可用，程序将无法正常运行")
    print("请安装必要的依赖包: pip install pandas openpyxl numpy")
elif not all(core_libraries):
    logger.warning("部分核心库不可用，某些功能可能受限")
    print("警告: 部分核心库不可用，某些功能可能受限")
    print("建议安装完整依赖: pip install pandas openpyxl sqlalchemy jinja2 reportlab requests schedule")

# 配置日志
import logging.handlers

# 创建日志记录器
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 创建控制台处理器
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# 创建文件处理器（带轮换功能）
# 获取程序所在目录
app_dir = Path(__file__).parent
log_file = app_dir / 'report_generator.log'
file_handler = logging.handlers.RotatingFileHandler(
    str(log_file),
    maxBytes=10*1024*1024,  # 10MB
    backupCount=5
)
file_handler.setLevel(logging.INFO)

# 创建安全日志处理器（记录敏感操作）
security_log_file = app_dir / 'security_report.log'
security_handler = logging.handlers.RotatingFileHandler(
    str(security_log_file),
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3
)
security_handler.setLevel(logging.WARNING)

# 定义日志格式
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)
security_handler.setFormatter(formatter)

# 添加处理器到日志记录器
logger.addHandler(console_handler)
logger.addHandler(file_handler)
logger.addHandler(security_handler)

# 创建安全日志记录器
security_logger = logging.getLogger('security')
security_logger.setLevel(logging.INFO)
security_logger.addHandler(security_handler)

class ConfigManager:
    """配置管理类，支持从环境变量和配置文件加载配置"""
    
    def __init__(self, config_file: Optional[str] = None, encrypted: bool = False):
        self.config = {}
        
        # 加载默认配置
        self._load_default_config()
        
        # 加载配置文件
        if config_file and os.path.exists(config_file):
            self._load_from_file(config_file, encrypted)
        
        # 加载环境变量
        self._load_from_env()
        
        # 验证配置
        self.validate_config()

    def _load_default_config(self):
        """加载默认配置"""
        self.config = {
            'output_dir': 'reports',
            'default_format': 'excel',
            'email': {
                'smtp_server': 'smtp.example.com',
                'smtp_port': 587,
                'use_tls': True
            },
            'logging': {
                'level': 'INFO',
                'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            }
        }

    def _load_from_file(self, config_file: str, encrypted: bool = False):
        """从配置文件加载配置（支持JSON和YAML，可选加密）"""
        try:
            file_ext = os.path.splitext(config_file)[1].lower()
            
            # 读取文件内容
            with open(config_file, 'rb' if encrypted else 'r', encoding='utf-8' if not encrypted else None) as f:
                content = f.read()
            
            # 如果是加密文件，先解密
            if encrypted:
                # 假设密钥存储在环境变量中，或者可以通过其他安全方式获取
                key = os.getenv('REPORT_CONFIG_KEY', None)
                if not key:
                    raise ValueError("加载加密配置文件需要设置环境变量 REPORT_CONFIG_KEY")
                content = SecurityUtils.decrypt_data(content, key)
            
            # 解析配置
            if file_ext == '.json':
                file_config = json.loads(content) if encrypted else json.loads(content.decode('utf-8'))
            elif file_ext in ['.yaml', '.yml']:
                try:
                    import yaml
                    file_config = yaml.safe_load(content) if encrypted else yaml.safe_load(content.decode('utf-8'))
                except ImportError:
                    logger.error("YAML配置文件需要PyYAML库，请安装: pip install pyyaml")
                    return
            else:
                logger.error(f"不支持的配置文件格式: {file_ext}")
                return
            
            self._merge_config(self.config, file_config)
            logger.info(f"成功从配置文件加载配置: {config_file}")
        except Exception as e:
            logger.error(f"加载配置文件失败: {e}")
    
    def _load_from_env(self):
        """从环境变量加载配置"""
        # 加载邮件配置
        self.config['email']['smtp_server'] = os.getenv('REPORT_SMTP_SERVER', self.config['email']['smtp_server'])
        self.config['email']['smtp_port'] = int(os.getenv('REPORT_SMTP_PORT', self.config['email']['smtp_port']))
        self.config['email']['username'] = os.getenv('REPORT_EMAIL_USERNAME', self.config['email'].get('username', ''))
        self.config['email']['password'] = os.getenv('REPORT_EMAIL_PASSWORD', self.config['email'].get('password', ''))
        
        # 加载其他配置
        self.config['output_dir'] = os.getenv('REPORT_OUTPUT_DIR', self.config['output_dir'])
    
    def _merge_config(self, dest: Dict[str, Any], src: Dict[str, Any]):
        """递归合并配置"""
        for key, value in src.items():
            if key in dest and isinstance(dest[key], dict) and isinstance(value, dict):
                self._merge_config(dest[key], value)
            else:
                dest[key] = value
    
    def validate_config(self):
        """验证配置的完整性和有效性"""
        errors = []
        warnings = []
        
        # 验证输出目录
        output_dir = self.config.get('output_dir')
        if output_dir:
            try:
                os.makedirs(output_dir, exist_ok=True)
            except OSError as e:
                errors.append(f"输出目录 '{output_dir}' 无法创建: {e}")
        else:
            errors.append("输出目录未配置")
        
        # 验证邮件配置
        email_config = self.config.get('email', {})
        if email_config.get('username') or email_config.get('password'):
            if not email_config.get('smtp_server'):
                errors.append("邮件配置中缺少SMTP服务器地址")
            if not email_config.get('smtp_port'):
                errors.append("邮件配置中缺少SMTP端口")
            else:
                try:
                    port = int(email_config['smtp_port'])
                    if port < 1 or port > 65535:
                        errors.append(f"邮件配置中SMTP端口 {port} 无效，必须在1-65535之间")
                except ValueError:
                    errors.append(f"邮件配置中SMTP端口 '{email_config['smtp_port']}' 不是有效的数字")
        
        # 验证日志配置
        logging_config = self.config.get('logging', {})
        if logging_config.get('level'):
            valid_levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL']
            if logging_config['level'].upper() not in valid_levels:
                warnings.append(f"日志级别 '{logging_config['level']}' 无效，将使用默认级别INFO")
                self.config['logging']['level'] = 'INFO'
        
        # 记录验证结果
        if errors:
            for error in errors:
                logger.error(error)
                print(f"配置错误: {error}")
            return False
        
        if warnings:
            for warning in warnings:
                logger.warning(warning)
                print(f"配置警告: {warning}")
        
        logger.info("配置验证通过")
        return True
    
    def get(self, key: str, default: Any = None) -> Any:
        """获取配置值"""
        keys = key.split('.')
        value = self.config
        
        for k in keys:
            if k in value:
                value = value[k]
            else:
                return default
        
        return value

# 创建全局配置管理器
class SecurityUtils:
    """安全工具类，用于处理加密和解密操作"""
    
    @staticmethod
    def generate_key(password: str, salt: bytes = None) -> bytes:
        """从密码生成加密密钥"""
        if salt is None:
            salt = os.urandom(16)
        
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        
        key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
        return key, salt
    
    @staticmethod
    def encrypt_data(data: Union[str, bytes], key: bytes) -> str:
        """加密数据"""
        try:
            f = Fernet(key)
            # 确保数据是字节类型
            if isinstance(data, str):
                data_bytes = data.encode()
            else:
                data_bytes = data
            encrypted_data = f.encrypt(data_bytes)
            result = base64.urlsafe_b64encode(encrypted_data).decode()
            security_logger.info("成功加密数据")
            return result
        except Exception as e:
            security_logger.error(f"加密数据失败: {e}")
            raise
    
    @staticmethod
    def decrypt_data(encrypted_data: Union[str, bytes], key: bytes) -> str:
        """解密数据"""
        try:
            f = Fernet(key)
            # 确保加密数据是字节类型
            if isinstance(encrypted_data, str):
                encrypted_bytes = base64.urlsafe_b64decode(encrypted_data.encode())
            else:
                encrypted_bytes = base64.urlsafe_b64decode(encrypted_data)
            result = f.decrypt(encrypted_bytes).decode()
            security_logger.info("成功解密数据")
            return result
        except Exception as e:
            security_logger.error(f"解密数据失败: {e}")
            raise
    
    @staticmethod
    def encrypt_config(config_dict: Dict[str, Any], password: str) -> Dict[str, Any]:
        """加密配置字典中的敏感字段"""
        key, salt = SecurityUtils.generate_key(password)
        
        # 定义需要加密的敏感字段
        sensitive_fields = [
            ('recipients', lambda x: True),  # 所有邮箱都加密
            ('parameters', lambda x: isinstance(x, dict) and 'password' in x),  # 参数中的密码
            ('data_source_path', lambda x: x and any(keyword in x.lower() for keyword in ['password', 'secret', 'token'])),
            ('schedule', lambda x: False),  # 不加密调度信息
        ]
        
        # 递归加密敏感字段
        def encrypt_recursive(obj, path=""):
            if isinstance(obj, dict):
                for key, value in obj.items():
                    new_path = f"{path}.{key}" if path else key
                    obj[key] = encrypt_recursive(value, new_path)
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    new_path = f"{path}[{i}]"
                    obj[i] = encrypt_recursive(item, new_path)
            else:
                # 检查是否需要加密
                for field, condition in sensitive_fields:
                    if field in path and condition(obj):
                        return {
                            "__encrypted__": True,
                            "value": SecurityUtils.encrypt_data(str(obj), key)
                        }
            return obj
        
        encrypted_config = encrypt_recursive(config_dict.copy())
        
        # 添加加密信息
        encrypted_config["__encryption__"] = {
            "version": "1.0",
            "salt": base64.urlsafe_b64encode(salt).decode()
        }
        
        return encrypted_config
    
    @staticmethod
    def decrypt_config(encrypted_config: Dict[str, Any], password: str) -> Dict[str, Any]:
        """解密配置字典中的敏感字段"""
        # 检查是否是加密配置
        if "__encryption__" not in encrypted_config:
            return encrypted_config
        
        # 提取加密信息
        encryption_info = encrypted_config["__encryption__"]
        salt = base64.urlsafe_b64decode(encryption_info["salt"].encode())
        
        # 生成密钥
        key, _ = SecurityUtils.generate_key(password, salt)
        
        # 递归解密敏感字段
        def decrypt_recursive(obj):
            if isinstance(obj, dict):
                if "__encrypted__" in obj:
                    return SecurityUtils.decrypt_data(obj["value"], key)
                
                for key, value in obj.items():
                    obj[key] = decrypt_recursive(value)
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    obj[i] = decrypt_recursive(item)
            return obj
        
        decrypted_config = decrypt_recursive(encrypted_config.copy())
        
        # 移除加密信息
        if "__encryption__" in decrypted_config:
            del decrypted_config["__encryption__"]
        
        return decrypted_config

# 权限管理模块


class ScheduleManager:
    """报表调度管理器"""
    def __init__(self):
        self.tasks = {}  # 存储任务：{task_id: (report_config, job)}，job是schedule库的任务对象
        self.running = False
        self._thread = None
        self.scheduler = schedule  # 使用schedule库的调度器
    
    def add_task(self, report_name: str, report_config: 'ReportConfig', schedule: str, email_recipients: List[str]) -> str:
        """添加调度任务
        
        Args:
            report_name: 报表名称
            report_config: 报表配置对象
            schedule: cron表达式
            email_recipients: 邮件接收者列表
        
        Returns:
            str: 任务ID
        """
        # 更新报表配置中的调度信息
        report_config.schedule = schedule
        report_config.email_recipients = email_recipients
        if not report_config.schedule:
            raise ValueError("报表配置中缺少调度表达式")
        
        # 生成唯一任务ID
        task_id = f"{report_config.report_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # 定义任务执行函数
        def job():
            try:
                logger.info(f"执行调度任务: {report_config.report_name}")
                engine = AutoReportEngine(report_config)
                result = engine.run()
                logger.info(f"调度任务完成: {report_config.report_name}，结果: {result}")
            except Exception as e:
                logger.error(f"调度任务执行失败: {report_config.report_name}，错误: {e}")
        
        # 根据cron表达式设置调度
        try:
            # 解析cron表达式 (分 时 日 月 周)
            parts = report_config.schedule.split()
            if len(parts) != 5:
                raise ValueError("无效的cron表达式格式，应为：分 时 日 月 周")
            
            minute, hour, day, month, weekday = parts
            
            # 使用schedule库添加任务
            # 注意：schedule库不直接支持cron表达式，这里使用简化实现
            # 仅支持每分钟、每小时和每天的调度
            if minute == "*":
                # 每分钟执行
                job = self.scheduler.every().minute.do(job)
            elif hour == "*":
                # 每小时执行
                job = self.scheduler.every().hour.at(f":{minute}").do(job)
            else:
                # 每天执行
                job = self.scheduler.every().day.at(f"{hour}:{minute}").do(job)
            
            self.tasks[task_id] = (report_config, job)
            logger.info(f"添加调度任务成功: {task_id}，表达式: {report_config.schedule}")
            return task_id
        except Exception as e:
            logger.error(f"添加调度任务失败: {task_id}，错误: {e}")
            raise
    
    def remove_task(self, task_id: str) -> None:
        """删除调度任务
        
        Args:
            task_id: 任务ID
        """
        if task_id not in self.tasks:
            raise ValueError(f"调度任务 {task_id} 不存在")
        
        report_config, job = self.tasks[task_id]
        
        # 删除任务
        self.scheduler.cancel_job(job)
        del self.tasks[task_id]
        logger.info(f"删除调度任务成功: {task_id}")
    
    def get_task(self, task_id: str) -> Optional['ReportConfig']:
        """获取调度任务信息
        
        Args:
            task_id: 任务ID
        
        Returns:
            Optional[ReportConfig]: 报表配置对象
        """
        if task_id not in self.tasks:
            return None
        
        return self.tasks[task_id][0]
    
    def get_all_tasks(self) -> List[Dict[str, Any]]:
        """获取所有调度任务
        
        Returns:
            List[Dict[str, Any]]: 任务列表，每个任务包含task_id、report_name、schedule、running等信息
        """
        return [{
            'task_id': task_id,
            'report_name': config.report_name,
            'schedule': config.schedule,
            'running': self.running
        } for task_id, (config, _) in self.tasks.items()]
    
    def start_scheduler(self) -> bool:
        """启动调度器
        
        Returns:
            bool: 是否成功启动
        """
        if self.running:
            logger.warning("调度器已经在运行")
            return False
        
        self.running = True
        logger.info("启动调度器")
        
        # 在单独的线程中运行调度器
        import threading
        self._thread = threading.Thread(target=self._run_scheduler, daemon=True)
        self._thread.start()
        return True
    
    def _run_scheduler(self):
        """在后台运行调度器的内部方法"""
        while self.running:
            self.scheduler.run_pending()
            time.sleep(1)
    
    def stop_scheduler(self) -> bool:
        """停止调度器
        
        Returns:
            bool: 是否成功停止
        """
        if not self.running:
            logger.warning("调度器已经停止")
            return False
        
        self.running = False
        if self._thread:
            self._thread.join(timeout=5)  # 等待线程结束，最多5秒
        
        # 清空所有任务
        for task_id in list(self.tasks.keys()):
            self.remove_task(task_id)
        
        logger.info("调度器已停止")
        return True
    
    def is_running(self) -> bool:
        """检查调度器是否正在运行
        
        Returns:
            bool: 调度器运行状态
        """
        return self.running

# 创建全局调度管理器
schedule_manager = ScheduleManager()

config_manager = ConfigManager()

@dataclass
class DataSourceConfig:
    """数据源配置类"""
    type: str  # 'excel', 'csv', 'sql', 'api'
    path: str  # 文件路径或连接字符串
    name: Optional[str] = None  # 数据源名称，用于标识
    parameters: Dict[str, Any] = field(default_factory=dict)  # 额外参数

@dataclass
class ReportConfig:
    """报表配置类"""
    report_name: str
    output_format: List[str]  # 输出格式：excel, pdf, html, email, csv
    data_sources: List[DataSourceConfig] = field(default_factory=list)  # 多数据源配置
    schedule: Optional[str] = None  # 调度表达式（如：0 0 * * *）
    recipients: List[str] = field(default_factory=list)  # 邮件接收者
    template_path: Optional[str] = None  # 报表模板路径
    template_type: Optional[str] = None  # 报表模板类型（如：default, simple, detailed, business）
    filters: Dict[str, Any] = field(default_factory=dict)  # 数据过滤条件
    calculations: List[Dict[str, str]] = field(default_factory=list)  # 计算字段
    charts: List[Dict[str, Any]] = field(default_factory=list)  # 图表配置
    parameters: Optional[Dict[str, Any]] = None  # 其他参数
    # 兼容旧版本的字段
    data_source_type: Optional[str] = None  # 兼容旧版本
    data_source_path: Optional[str] = None  # 兼容旧版本


def load_config_from_file(file_path: str, encrypted: bool = False) -> Dict[str, Any]:
    """加载配置文件（支持JSON和YAML，可选加密）"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # 读取文件内容
        with open(file_path, 'rb' if encrypted else 'r', encoding='utf-8' if not encrypted else None) as f:
            content = f.read()
        
        # 如果是加密文件，先解密
        if encrypted:
            # 假设密钥存储在环境变量中，或者可以通过其他安全方式获取
            key = os.getenv('REPORT_CONFIG_KEY', None)
            if not key:
                raise ValueError("加载加密配置文件需要设置环境变量 REPORT_CONFIG_KEY")
            content = SecurityUtils.decrypt_data(content, key)
        
        # 解析配置
        if file_ext == '.json':
            return json.loads(content) if encrypted else json.loads(content.decode('utf-8'))
        elif file_ext in ['.yaml', '.yml']:
            try:
                import yaml
                return yaml.safe_load(content) if encrypted else yaml.safe_load(content.decode('utf-8'))
            except ImportError:
                logger.error("YAML配置文件需要PyYAML库，请安装: pip install pyyaml")
                raise
        else:
            raise ValueError(f"不支持的配置文件格式: {file_ext}")
    except Exception as e:
        logger.error(f"加载配置文件失败: {e}")
        raise


def save_config_to_file(config: Dict[str, Any], file_path: str, encrypted: bool = False) -> None:
    """保存配置到文件（支持JSON和YAML，可选加密）"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # 生成配置内容
        if file_ext == '.json':
            content = json.dumps(config, ensure_ascii=False, indent=2)
        elif file_ext in ['.yaml', '.yml']:
            try:
                import yaml
                content = yaml.dump(config, default_flow_style=False, allow_unicode=True)
            except ImportError:
                logger.error("YAML配置文件需要PyYAML库，请安装: pip install pyyaml")
                raise
        else:
            raise ValueError(f"不支持的配置文件格式: {file_ext}")
        
        # 如果需要加密，先加密内容
        if encrypted:
            # 假设密钥存储在环境变量中，或者可以通过其他安全方式获取
            key = os.getenv('REPORT_CONFIG_KEY', None)
            if not key:
                raise ValueError("保存加密配置文件需要设置环境变量 REPORT_CONFIG_KEY")
            content = SecurityUtils.encrypt_data(content, key)
        
        # 写入文件
        with open(file_path, 'wb' if encrypted else 'w', encoding='utf-8' if not encrypted else None) as f:
            f.write(content)
        
        logger.info(f"配置已保存到文件: {file_path}")
    except Exception as e:
        logger.error(f"保存配置文件失败: {e}")
        raise


def validate_report_config(config: Dict[str, Any]) -> bool:
    """验证报表配置"""
    required_fields = ['report_name', 'data_source_type', 'data_source_path', 'output_format']
    
    for field in required_fields:
        if field not in config:
            logger.error(f"缺少必要的配置字段: {field}")
            return False
    
    valid_data_source_types = ['excel', 'csv', 'sql', 'api']
    if config['data_source_type'] not in valid_data_source_types:
        logger.error(f"无效的数据源类型: {config['data_source_type']}")
        return False
    
    valid_output_formats = ['excel', 'pdf', 'html', 'email']
    for fmt in config['output_format']:
        if fmt not in valid_output_formats:
            logger.error(f"无效的输出格式: {fmt}")
            return False
    
    return True


@dataclass
class DataSourceConfig:
    """数据源配置类"""
    type: str  # 'excel', 'csv', 'sql', 'api'
    path: str  # 文件路径或连接字符串
    name: Optional[str] = None  # 数据源名称，用于标识
    parameters: Dict[str, Any] = field(default_factory=dict)  # 额外参数

def get_config_from_dict(config_dict: Dict[str, Any]) -> ReportConfig:
    """从字典创建ReportConfig实例"""
    # 处理多数据源配置
    data_sources = []
    
    # 兼容旧版本配置
    if 'data_source_type' in config_dict and 'data_source_path' in config_dict:
        data_sources.append(DataSourceConfig(
            type=config_dict['data_source_type'],
            path=config_dict['data_source_path'],
            name='default',
            parameters=config_dict.get('parameters', {})
        ))
    
    # 处理新版本多数据源配置
    if 'data_sources' in config_dict:
        for ds_config in config_dict['data_sources']:
            data_sources.append(DataSourceConfig(**ds_config))
    
    return ReportConfig(
        report_name=config_dict['report_name'],
        data_sources=data_sources,
        output_format=config_dict['output_format'],
        schedule=config_dict.get('schedule'),
        recipients=config_dict.get('recipients', []),
        template_path=config_dict.get('template_path'),
        filters=config_dict.get('filters', {}),
        calculations=config_dict.get('calculations', []),
        charts=config_dict.get('charts', []),
        parameters=config_dict.get('parameters', {}),
        # 兼容旧版本的字段
        data_source_type=config_dict.get('data_source_type'),
        data_source_path=config_dict.get('data_source_path')
    )

class DataSource(ABC):
    """数据源抽象基类"""
    
    @abstractmethod
    def load_data(self, config: ReportConfig) -> 'pd.DataFrame':
        pass
    
    def validate_data(self, df: 'pd.DataFrame') -> bool:
        """验证数据（默认实现）"""
        # 基础数据验证
        if df.empty:
            logger.warning("数据框为空")
            return False
        return True

class ExcelDataSource(DataSource):
    """Excel数据源"""
    
    def load_data(self, config: ReportConfig) -> 'pd.DataFrame':
        try:
            logger.info(f"从Excel加载数据: {config.data_source_path}")
            
            # 支持多种Excel读取方式
            if config.data_source_path.endswith('.xlsx'):
                engine = 'openpyxl'
            elif config.data_source_path.endswith('.xls'):
                engine = 'xlrd'
            else:
                engine = None
            
            # 获取分块大小参数
            chunksize = config.parameters.get('chunksize', None)
            
            # 读取Excel文件
            if 'sheet_name' in config.parameters:
                sheet_name = config.parameters.get('sheet_name')
                
                if chunksize:
                    # 分块加载大数据
                    chunks = []
                    for chunk in pd.read_excel(
                        config.data_source_path,
                        engine=engine,
                        sheet_name=sheet_name,
                        chunksize=chunksize
                    ):
                        chunks.append(chunk)
                    df = pd.concat(chunks, ignore_index=True)
                else:
                    df = pd.read_excel(
                        config.data_source_path,
                        engine=engine,
                        sheet_name=sheet_name
                    )
            else:
                # 尝试读取所有sheet
                xl = pd.ExcelFile(config.data_source_path, engine=engine)
                if len(xl.sheet_names) > 1:
                    dfs = {}
                    for sheet in xl.sheet_names:
                        if chunksize:
                            chunks = []
                            for chunk in xl.parse(sheet, chunksize=chunksize):
                                chunks.append(chunk)
                            dfs[sheet] = pd.concat(chunks, ignore_index=True)
                        else:
                            dfs[sheet] = xl.parse(sheet)
                    # 合并所有sheet或选择第一个
                    df = dfs[list(dfs.keys())[0]]
                else:
                    sheet_name = xl.sheet_names[0]
                    if chunksize:
                        chunks = []
                        for chunk in xl.parse(sheet_name, chunksize=chunksize):
                            chunks.append(chunk)
                        df = pd.concat(chunks, ignore_index=True)
                    else:
                        df = xl.parse(sheet_name)
            
            logger.info(f"数据加载完成，形状: {df.shape}")
            return df
            
        except FileNotFoundError:
            logger.error(f"Excel文件不存在: {config.data_source_path}")
            raise
        except PermissionError:
            logger.error(f"没有权限读取Excel文件: {config.data_source_path}")
            raise
        except pd.errors.EmptyDataError:
            logger.error(f"Excel文件为空: {config.data_source_path}")
            raise
        except pd.errors.ParserError as e:
            logger.error(f"Excel文件解析错误: {e}")
            raise
        except Exception as e:
            logger.error(f"Excel数据加载失败: {e}")
            raise
    
    def validate_data(self, df: 'pd.DataFrame') -> bool:
        # 基础数据验证
        if df.empty:
            logger.warning("数据框为空")
            return False
        
        # 检查必要列
        required_cols = ['日期', '金额', '类别']  # 可根据配置调整
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            logger.warning(f"缺少必要列: {missing_cols}")
            return False
        
        return True

class CSVDataSource(DataSource):
    """CSV数据源"""
    
    def load_data(self, config: ReportConfig) -> 'pd.DataFrame':
        try:
            logger.info(f"从CSV加载数据: {config.data_source_path}")
            
            # 检查文件大小
            file_size = os.path.getsize(config.data_source_path)
            
            # 设置默认参数，自动解析日期列
            csv_params = config.parameters.copy()
            if 'parse_dates' not in csv_params:
                # 尝试自动解析所有可能的日期列
                csv_params['parse_dates'] = True
            
            if file_size > 100 * 1024 * 1024:  # 大于100MB
                logger.info(f"检测到大型CSV文件 ({file_size/1024/1024:.2f}MB)，使用分块加载")
                
                # 分块读取数据
                chunks = []
                for chunk in pd.read_csv(
                    config.data_source_path,
                    chunksize=100000,  # 10万行/块
                    **csv_params
                ):
                    chunks.append(chunk)
                    logger.info(f"已加载 {len(chunks) * 100000} 行数据")
                
                # 合并数据
                df = pd.concat(chunks, ignore_index=True)
            else:
                # 小文件直接读取
                df = pd.read_csv(config.data_source_path, **csv_params)
            
            logger.info(f"成功加载 {len(df)} 行数据")
            return df
        except Exception as e:
            logger.error(f"加载CSV数据失败: {e}")
            raise

class SQLDataSource(DataSource):
    """SQL数据库数据源"""
    
    def load_data(self, config: ReportConfig) -> 'pd.DataFrame':
        try:
            logger.info(f"从SQL数据库加载数据")
            
            # 构建数据库连接
            connection_str = config.data_source_path
            query = config.parameters.get('query')
            params = config.parameters.get('params', {})
            
            if not query:
                raise ValueError("SQL查询不能为空")
            
            engine = sa.create_engine(connection_str)
            
            # 执行查询，支持参数化查询
            with engine.connect() as conn:
                df = pd.read_sql_query(query, conn, params=params)
            
            logger.info(f"数据加载完成，形状: {df.shape}")
            return df
            
        except ValueError as e:
            logger.error(f"SQL参数错误: {e}")
            raise
        except sa.exc.OperationalError as e:
            logger.error(f"数据库连接错误: {e}")
            raise
        except sa.exc.ProgrammingError as e:
            logger.error(f"SQL查询语法错误: {e}")
            raise
        except pd.errors.DatabaseError as e:
            logger.error(f"数据库操作错误: {e}")
            raise
        except Exception as e:
            logger.error(f"SQL数据加载失败: {e}")
            raise
    
    def validate_data(self, df: 'pd.DataFrame') -> bool:
        return not df.empty


class APIDataSource(DataSource):
    """API数据源"""
    
    def load_data(self, config: ReportConfig) -> 'pd.DataFrame':
        """从API加载数据"""
        try:
            # 获取API配置参数
            url = config.data_source_path
            method = config.parameters.get('method', 'GET').upper()
            headers = config.parameters.get('headers', {})
            params = config.parameters.get('params', {})
            data = config.parameters.get('data', {})
            json_data = config.parameters.get('json', None)
            auth = config.parameters.get('auth', None)
            response_key = config.parameters.get('response_key', None)
            timeout = config.parameters.get('timeout', 30)
            retries = config.parameters.get('retries', 3)
            retry_delay = config.parameters.get('retry_delay', 1)
            
            # 发送请求（带重试机制）
            for attempt in range(retries):
                try:
                    logger.info(f"发送{method}请求到API: {url} (尝试 {attempt+1}/{retries})")
                    
                    if method == 'GET':
                        response = requests.get(url, headers=headers, params=params, auth=auth, timeout=timeout)
                    elif method == 'POST':
                        response = requests.post(url, headers=headers, params=params, 
                                               data=data, json=json_data, auth=auth, timeout=timeout)
                    elif method == 'PUT':
                        response = requests.put(url, headers=headers, params=params, 
                                              data=data, json=json_data, auth=auth, timeout=timeout)
                    elif method == 'DELETE':
                        response = requests.delete(url, headers=headers, params=params, auth=auth, timeout=timeout)
                    else:
                        raise ValueError(f"不支持的HTTP方法: {method}")
                    
                    # 检查响应状态
                    response.raise_for_status()
                    break  # 成功，退出重试循环
                
                except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                    if attempt < retries - 1:
                        logger.warning(f"API请求失败，将重试: {e}")
                        import time
                        time.sleep(retry_delay)
                    else:
                        logger.error(f"API请求失败，已达到最大重试次数: {e}")
                        raise
            
            # 解析响应
            response_data = response.json()
            
            # 如果指定了响应键，提取对应的数据
            if response_key:
                for key in response_key.split('.'):
                    response_data = response_data.get(key, {})
            
            # 转换为DataFrame
            df = pd.DataFrame(response_data)
            
            logger.info(f"从API加载了 {len(df)} 条记录")
            return df
            
        except requests.exceptions.HTTPError as e:
            logger.error(f"API请求失败 (HTTP错误): {e}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"API响应解析错误: {e}")
            raise
        except ValueError as e:
            logger.error(f"API参数错误: {e}")
            raise
        except Exception as e:
            logger.error(f"从API加载数据失败: {e}")
            raise

class DataProcessor:
    """数据处理类"""
    
    @staticmethod
    def apply_filters(df: 'pd.DataFrame', filters: Dict[str, Union[Dict[str, Any], List[Any], Any]]) -> 'pd.DataFrame':
        """应用数据过滤器"""
        if not filters:
            return df
        
        try:
            # 创建一个掩码数组而不是多次复制数据框
            mask = pd.Series([True] * len(df), index=df.index)
            
            for column, condition in filters.items():
                # 检查列是否存在
                if column not in df.columns:
                    logger.warning(f"过滤列 '{column}' 不存在于数据中，跳过该过滤条件")
                    continue
                    
                if isinstance(condition, dict):
                    # 范围过滤
                    if 'min' in condition and 'max' in condition:
                        try:
                            mask &= (df[column] >= condition['min']) & (df[column] <= condition['max'])
                        except TypeError as e:
                            logger.warning(f"范围过滤失败，列 '{column}' 数据类型不匹配: {e}")
                    # 值列表过滤
                    elif 'values' in condition:
                        mask &= df[column].isin(condition['values'])
                elif isinstance(condition, list):
                    # 多值过滤
                    mask &= df[column].isin(condition)
                else:
                    # 单值过滤
                    try:
                        mask &= (df[column] == condition)
                    except TypeError as e:
                        logger.warning(f"单值过滤失败，列 '{column}' 数据类型不匹配: {e}")
            
            # 应用掩码
            filtered_df = df[mask].copy()
            
            # 检查过滤后的数据是否为空
            if filtered_df.empty:
                logger.warning("应用过滤器后数据为空")
            
            return filtered_df
        except Exception as e:
            logger.error(f"应用过滤器失败: {e}")
            # 如果过滤失败，返回原始数据而不是抛出异常
            return df
    
    @staticmethod
    def apply_calculations(df: 'pd.DataFrame', calculations: List[Dict[str, Any]]) -> 'pd.DataFrame':
        """应用计算字段"""
        if not calculations:
            return df
        
        result_df = df.copy()
        
        for calc in calculations:
            column_name = calc.get('column')
            formula = calc.get('formula')
            operation = calc.get('operation')
            
            if column_name and formula:
                try:
                    # 简化执行环境，使用更安全的方式执行计算
                    result_df[column_name] = df.eval(formula, engine='python')
                    logger.info(f"成功计算字段: {column_name}")
                except Exception as e:
                    logger.warning(f"计算失败: {formula}，错误: {str(e)}")
            elif operation == 'groupby':
                # 分组聚合
                group_cols = calc.get('group_by', [])
                agg_cols = calc.get('aggregate', {})
                
                if group_cols and agg_cols:
                    grouped = result_df.groupby(group_cols).agg(agg_cols).reset_index()
                    logger.info(f"成功执行分组聚合: {group_cols}")
                    return grouped
            elif operation == 'pivot':
                # 透视表
                index = calc.get('index', [])
                columns = calc.get('columns', [])
                values = calc.get('values', [])
                aggfunc = calc.get('aggfunc', 'mean')
                
                if index and values:
                    pivoted = pd.pivot_table(
                        result_df, 
                        index=index, 
                        columns=columns, 
                        values=values, 
                        aggfunc=aggfunc, 
                        fill_value=0
                    ).reset_index()
                    logger.info(f"成功执行透视表: {index} -> {values}")
                    return pivoted
        
        return result_df
    
    @staticmethod
    def preview_data(df: 'pd.DataFrame', num_rows: int = 10) -> str:
        """生成数据预览字符串"""
        return df.head(num_rows).to_string(max_columns=10, line_width=100)
    
    @staticmethod
    def export_data_preview(df: 'pd.DataFrame', output_path: str, num_rows: int = 100) -> str:
        """导出数据预览"""
        preview_df = df.head(num_rows)
        
        # 根据扩展名选择导出格式
        ext = Path(output_path).suffix.lower()
        
        if ext == '.xlsx':
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                preview_df.to_excel(writer, index=False, sheet_name='Preview')
        elif ext == '.csv':
            preview_df.to_csv(output_path, index=False)
        elif ext == '.html':
            preview_df.to_html(output_path, index=False)
        else:
            raise ValueError(f"不支持的预览格式: {ext}")
        
        logger.info(f"数据预览已导出: {output_path}")
        return output_path
    
    @staticmethod
    def merge_dataframes(dataframes: List['pd.DataFrame'], merge_config: Optional[Dict[str, Any]] = None) -> 'pd.DataFrame':
        """合并多个数据框
        
        Args:
            dataframes: 要合并的数据框列表
            merge_config: 合并配置
                - on: 合并的列名或列名列表
                - how: 合并方式 ('inner', 'outer', 'left', 'right'), 默认 'inner'
                - suffixes: 列名冲突时的后缀, 默认 ('_1', '_2')
                - merge_type: 'merge' 或 'concat', 默认 'merge'
                - axis: concat 时的轴, 默认 0
        
        Returns:
            合并后的数据框
        """
        if not dataframes:
            raise ValueError("没有数据框可合并")
            
        if len(dataframes) == 1:
            return dataframes[0]
            
        # 默认合并配置
        default_config = {
            'how': 'inner',
            'suffixes': ('_1', '_2'),
            'merge_type': 'merge',
            'axis': 0
        }
        
        if merge_config:
            default_config.update(merge_config)
        
        merge_type = default_config.pop('merge_type')
        
        try:
            if merge_type == 'merge':
                # 多表合并
                result = dataframes[0]
                for df in dataframes[1:]:
                    # 查找共同列作为合并键
                    common_cols = list(set(result.columns) & set(df.columns))
                    if common_cols:
                        on = default_config.get('on', common_cols[0])
                        
                        try:
                            # 尝试统一日期列的数据类型
                            if on in result.columns and on in df.columns:
                                # 检查是否都是日期时间类型
                                result_dtype = str(result[on].dtype)
                                df_dtype = str(df[on].dtype)
                                
                                # 如果类型不一致，尝试转换为datetime
                                if 'datetime' in result_dtype and 'datetime' not in df_dtype:
                                    df[on] = pd.to_datetime(df[on], errors='coerce')
                                elif 'datetime' not in result_dtype and 'datetime' in df_dtype:
                                    result[on] = pd.to_datetime(result[on], errors='coerce')
                                elif 'datetime' not in result_dtype and 'datetime' not in df_dtype:
                                    # 尝试将两者都转换为datetime（如果可能）
                                    try:
                                        result[on] = pd.to_datetime(result[on])
                                        df[on] = pd.to_datetime(df[on])
                                    except:
                                        # 如果转换失败，保持原样
                                        pass
                            
                            # 为pd.merge创建配置（移除axis参数，因为merge不支持）
                            merge_config = default_config.copy()
                            if 'axis' in merge_config:
                                del merge_config['axis']
                            result = pd.merge(result, df, on=on, **merge_config)
                            logger.info(f"合并完成，当前数据框形状: {result.shape}")
                        except Exception as e:
                            # 合并失败，使用concat代替
                            logger.warning(f"合并失败: {e}，使用concat进行合并")
                            # 为pd.concat创建配置，使用axis=1并移除pd.concat不支持的参数
                            concat_config = default_config.copy()
                            concat_config['axis'] = 1
                            # 移除concat不支持的参数
                            for param in ['how', 'on', 'suffixes']:
                                if param in concat_config:
                                    del concat_config[param]
                            result = pd.concat([result, df], **concat_config)
                    else:
                        # 没有共同列，使用 concat
                        logger.warning("没有找到共同列，使用 concat 进行合并")
                        # 为pd.concat创建配置，使用axis=1并移除pd.concat不支持的参数
                        concat_config = default_config.copy()
                        concat_config['axis'] = 1
                        # 移除concat不支持的参数
                        for param in ['how', 'on', 'suffixes']:
                            if param in concat_config:
                                del concat_config[param]
                        result = pd.concat([result, df], **concat_config)
            else:
                # 简单拼接
                # 确保concat配置正确，移除pd.concat不支持的参数
                concat_config = default_config.copy()
                # 移除concat不支持的参数
                for param in ['how', 'on', 'suffixes']:
                    if param in concat_config:
                        del concat_config[param]
                result = pd.concat(dataframes, **concat_config)
                logger.info(f"拼接完成，数据框形状: {result.shape}")
            
            return result
        except Exception as e:
            logger.error(f"数据框合并失败: {e}")
            raise
    
    @staticmethod
    def calculate_metrics(df: 'pd.DataFrame') -> Dict[str, Union[int, Dict[str, Any]]]:
        """计算关键指标（优化版）"""
        metrics = {
            'total_records': len(df),
            'total_columns': len(df.columns),
            'missing_values': {},
            'numeric_stats': {},
            'categorical_stats': {}
        }
        
        # 计算缺失值（优化：使用更高效的方式）
        missing_values = df.isnull().sum()
        metrics['missing_values'] = missing_values[missing_values > 0].to_dict()
        
        # 数值列统计
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            # 对于大型数据集，只计算基本统计量以提高性能
            if len(df) > 100000:
                # 使用describe()一次性计算所有基本统计量，比逐个计算更快
                desc_stats = df[numeric_cols].describe()
                for col in numeric_cols:
                    metrics['numeric_stats'][col] = {
                        'mean': desc_stats[col]['mean'],
                        'min': desc_stats[col]['min'],
                        'max': desc_stats[col]['max'],
                        'std': desc_stats[col]['std']
                        # 对于大数据集，不计算中位数以提高性能
                    }
            else:
                # 小数据集计算完整统计量
                for col in numeric_cols:
                    metrics['numeric_stats'][col] = {
                        'mean': df[col].mean(),
                        'median': df[col].median(),
                        'min': df[col].min(),
                        'max': df[col].max(),
                        'std': df[col].std()
                    }
        
        # 分类列统计
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            for col in categorical_cols:
                # 优化：只计算非空值的唯一数
                unique_count = df[col].nunique(dropna=True)
                
                # 对于大型数据集，不计算top_value以提高性能
                if len(df) > 100000:
                    metrics['categorical_stats'][col] = {
                        'unique_count': unique_count
                    }
                else:
                    # 小数据集计算完整统计量
                    top_value = df[col].mode().iloc[0] if not df[col].mode().empty else None
                    metrics['categorical_stats'][col] = {
                        'unique_count': unique_count,
                        'top_value': top_value
                    }
        
        return metrics

class ReportGenerator(ABC):
    """报表生成器抽象基类"""
    
    @abstractmethod
    def generate(self, df: 'pd.DataFrame', metrics: Dict[str, Any], output_path: str, charts: Optional[List[Dict[str, Any]]] = None) -> str:
        """生成报表"""
        pass

# Excel报表生成器优化
class ExcelReportGenerator(ReportGenerator):
    """Excel报表生成器（优化版）"""
    
    def generate(self, df: 'pd.DataFrame', metrics: Dict[str, Any], output_path: str, charts: Optional[List[Dict[str, Any]]] = None) -> str:
        try:
            logger.info(f"生成Excel报表: {output_path}")
            
            # 对于大型数据集，只写入前100万行
            if len(df) > 1000000:
                logger.info(f"数据集过大 ({len(df)} 行)，只写入前100万行")
                df_to_write = df.head(1000000)
            else:
                df_to_write = df
            
            # 创建Excel工作簿
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 写入原始数据
                df_to_write.to_excel(writer, sheet_name='数据', index=False)
                
                # 创建摘要表
                summary_df = self._create_summary_table(df, metrics)
                summary_df.to_excel(writer, sheet_name='摘要', index=False)
                
                # 获取工作表
                data_worksheet = writer.sheets['数据']
                summary_worksheet = writer.sheets['摘要']
                
                # 应用样式
                self._apply_excel_styles(data_worksheet, df_to_write)
                self._apply_excel_styles(summary_worksheet, summary_df)
                
                # 自动调整列宽
                self._auto_adjust_columns(data_worksheet, df_to_write)
                self._auto_adjust_columns(summary_worksheet, summary_df)
                
                # 添加图表（仅当数据量适中时）
                if charts and len(df_to_write) < 10000:
                    self._add_charts(writer.book, df_to_write, charts)
            
            logger.info(f"Excel报表生成成功: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"生成Excel报表失败: {e}")
            raise
    
    def _create_summary_table(self, df: 'pd.DataFrame', metrics: Dict[str, Any]) -> 'pd.DataFrame':
        """创建摘要表"""
        summary_data = {
            '指标': [],
            '值': []
        }
        
        # 添加基本统计信息
        summary_data['指标'].extend(['总记录数', '总列数'])
        summary_data['值'].extend([metrics['total_records'], metrics['total_columns']])
        
        # 添加数值列统计
        for col, stats in metrics['numeric_stats'].items():
            for stat_name, value in stats.items():
                summary_data['指标'].append(f"{col} - {stat_name}")
                summary_data['值'].append(value)
        
        return pd.DataFrame(summary_data)
    
    def _apply_excel_styles(self, worksheet, df: 'pd.DataFrame'):
        """应用Excel样式"""
        # 简单的样式应用
        pass
    
    def _auto_adjust_columns(self, worksheet, df: 'pd.DataFrame'):
        """自动调整列宽"""
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    def _add_charts(self, workbook, df: 'pd.DataFrame', charts: List[Dict[str, Any]]):
        """添加图表"""
        # 简单的图表添加实现
        pass

# PDF报表生成器优化
class PDFReportGenerator(ReportGenerator):
    """PDF报表生成器（优化版）"""
    
    def generate(self, df: 'pd.DataFrame', metrics: Dict[str, Any], output_path: str, charts: Optional[List[Dict[str, Any]]] = None) -> str:
        try:
            logger.info(f"生成PDF报表: {output_path}")
            
            # 对于大型数据集，只显示前100行
            if len(df) > 100:
                logger.info(f"数据集过大 ({len(df)} 行)，PDF中只显示前100行")
                df_to_display = df.head(100)
            else:
                df_to_display = df
            
            # 创建PDF文档
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 添加标题
            title = Paragraph("自动化报表", styles['Heading1'])
            story.append(title)
            
            # 添加生成时间
            generated_time = Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            story.append(generated_time)
            
            # 添加空行
            story.append(Paragraph("", styles['Normal']))
            
            # 添加数据摘要
            summary_title = Paragraph("数据摘要", styles['Heading2'])
            story.append(summary_title)
            
            summary_data = [
                ['总记录数', str(metrics['total_records'])],
                ['总列数', str(metrics['total_columns'])]
            ]
            
            summary_table = Table(summary_data, colWidths=[100, 100])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            
            # 添加空行
            story.append(Paragraph("", styles['Normal']))
            
            # 添加数据表格
            data_title = Paragraph("数据内容", styles['Heading2'])
            story.append(data_title)
            
            # 转换DataFrame为列表
            data = [df_to_display.columns.tolist()] + df_to_display.values.tolist()
            
            # 创建表格
            data_table = Table(data, colWidths=[100] * len(df_to_display.columns))
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10)
            ]))
            story.append(data_table)
            
            # 生成PDF
            doc.build(story)
            
            logger.info(f"PDF报表生成成功: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"生成PDF报表失败: {e}")
            raise

# Excel报表生成器辅助方法
    def _create_summary_table(self, df: 'pd.DataFrame', metrics: Dict[str, Any]) -> 'pd.DataFrame':
        """创建摘要表"""
        summary_data = {
            '指标': [],
            '值': []
        }
        
        # 添加基本统计信息
        summary_data['指标'].extend(['总记录数', '总列数'])
        summary_data['值'].extend([metrics['total_records'], metrics['total_columns']])
        
        # 添加数值列统计
        for col, stats in metrics['numeric_stats'].items():
            for stat_name, value in stats.items():
                summary_data['指标'].append(f"{col} - {stat_name}")
                summary_data['值'].append(value)
        
        return pd.DataFrame(summary_data)
    
    def _apply_excel_styles(self, worksheet, df: 'pd.DataFrame'):
        """应用Excel样式"""
        # 简单的样式应用
        pass
    
    def _auto_adjust_columns(self, worksheet, df: 'pd.DataFrame'):
        """自动调整列宽"""
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
    
    def _add_charts(self, workbook, df: 'pd.DataFrame', charts: List[Dict[str, Any]]):
        """添加图表"""
        # 简单的图表添加实现
        pass

# HTML报表生成器优化
class HTMLReportGenerator(ReportGenerator):
    """HTML报表生成器（优化版）"""
    
    def __init__(self, template_path: Optional[str] = None, template_type: Optional[str] = None):
        self.template_path = template_path
        self.template_type = template_type
        self.templates_dir = os.path.join(os.getcwd(), 'templates')
    
    def get_available_templates(self) -> List[str]:
        """获取可用的模板列表"""
        if not os.path.exists(self.templates_dir):
            return []
        
        templates = []
        for filename in os.listdir(self.templates_dir):
            if filename.endswith(('.html', '.jinja2')):
                templates.append(os.path.splitext(filename)[0])
        return templates
    
    def _get_template_path(self) -> str:
        """获取模板文件路径"""
        if self.template_path:
            return self.template_path
        
        # 根据模板类型选择模板文件
        if self.template_type:
            template_file = f"{self.template_type}.html"
            template_path = os.path.join(self.templates_dir, template_file)
            if os.path.exists(template_path):
                return template_path
        
        # 如果没有指定模板或模板不存在，使用默认模板
        return None
    
    def _generate_from_template(self, df: 'pd.DataFrame', metrics: Dict[str, Any], charts: Optional[List[Dict[str, Any]]] = None, report_name: Optional[str] = None) -> str:
        """使用自定义模板生成HTML内容"""
        template_path = self._get_template_path()
        if not template_path:
            return self._generate_default_html(df, metrics, charts, report_name)
        
        try:
            # 读取模板文件
            with open(template_path, 'r', encoding='utf-8') as f:
                template_content = f.read()
            
            # 对于大型数据集，只显示前1000行
            large_data = len(df) > 1000
            if large_data:
                df_to_display = df.head(1000)
            else:
                df_to_display = df
            
            # 准备summary数据（将metrics转换为列表格式）
            summary = []
            for key, value in metrics.items():
                summary.append({'指标': key, '值': value})
            
            # 准备模板数据
            generated_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            template_data = {
                'report_name': report_name or '自动化报表',
                'generated_time': generated_time,
                'generation_time': generated_time,  # 兼容不同模板的时间变量名
                'total_rows': metrics['total_records'],
                'total_columns': metrics['total_columns'],
                'data_table': df_to_display.to_html(index=False),
                'large_data': large_data,
                'metrics': metrics,
                'summary': summary,
                'charts': charts or [],
                'data_source': '自动化报表工具',  # 添加数据源信息
                'total_records': metrics['total_records'],  # 确保total_records存在
                'total_columns': metrics['total_columns'],  # 确保total_columns存在
                'filters': {},  # 添加空的filters字典，防止模板引用时出错
                'has_data': not df.empty,  # 添加数据存在标志，避免直接判断DataFrame
                'data_columns': list(df.columns)  # 仅传递列名列表
            }
            
            # 使用Jinja2渲染模板（如果可用）
            if jinja2:
                template = jinja2.Template(template_content)
                return template.render(**template_data)
            else:
                # 简单的字符串替换作为回退
                html_content = template_content
                for key, value in template_data.items():
                    placeholder = f"{{{{ {key} }}}}"
                    if placeholder in html_content:
                        html_content = html_content.replace(placeholder, str(value))
                return html_content
        except Exception as e:
            logger.error(f"使用模板生成HTML失败: {e}")
            return self._generate_default_html(df, metrics, charts, report_name)
    
    def generate(self, df: 'pd.DataFrame', metrics: Dict[str, Any], output_path: str, charts: Optional[List[Dict[str, Any]]] = None, report_name: Optional[str] = None) -> str:
        try:
            logger.info(f"生成HTML报表: {output_path}")
            
            # 检查是否使用自定义模板
            if self.template_path or self.template_type:
                html_content = self._generate_from_template(df, metrics, charts, report_name)
            else:
                html_content = self._generate_default_html(df, metrics, charts, report_name)
            
            # 写入HTML文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"HTML报表生成成功: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"生成HTML报表失败: {e}")
            raise
    
    def _generate_default_html(self, df: 'pd.DataFrame', metrics: Dict[str, Any], charts: Optional[List[Dict[str, Any]]] = None, report_name: Optional[str] = None) -> str:
        """生成默认HTML报表（优化版）"""
        # 生成默认HTML报表（优化版）
        # 使用Jinja2语法的风格，但避免使用format方法处理CSS大括号
        
        # 生成CSS样式部分，避免使用format方法
        css_style = '''
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            line-height: 1.6;
        }
        h1 {
            color: #4F81BD;
            text-align: center;
        }
        h2 {
            color: #4F81BD;
            margin-top: 30px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            padding: 8px;
            text-align: left;
            border: 1px solid #ddd;
        }
        th {
            background-color: #4F81BD;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f2f2f2;
        }
        .summary {
            width: 50%;
        }
        .generated-time {
            text-align: right;
            font-style: italic;
            color: #666;
        }
        .data-table-container {
            overflow-x: auto;
        }
        .large-data-notice {
            background-color: #fff3cd;
            border: 1px solid #ffeaa7;
            padding: 10px;
            border-radius: 4px;
            margin: 10px 0;
        }
        .chart-container {
            margin: 20px 0;
            padding: 10px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }'''
        
        # 生成HTML主体部分，使用字符串连接而非format方法
        html_start = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{report_name or '自动化报表'}</title>
    <style>{css_style}</style>
</head>
<body>
    <h1>{report_name or '自动化报表'}</h1>
    <p class="generated-time">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    
    <h2>数据摘要</h2>
    <table class="summary">
        <tr>
            <th>项目</th>
            <th>值</th>
        </tr>
        <tr>
            <td>总记录数</td>
            <td>{metrics['total_records']}</td>
        </tr>
        <tr>
            <td>总列数</td>
            <td>{metrics['total_columns']}</td>
        </tr>
    </table>'''
        
        # 处理图表部分
        charts_section = ""
        if charts:
            charts_section = "\n    <h2>图表分析</h2>"
            for chart in charts:
                charts_section += f'''\n    <div class="chart-container">
        <h3>{chart.get('title', '图表')}</h3>
        <p>图表类型: {chart.get('type', '未指定')}</p>
        <p>数据范围: {chart.get('data_range', '未指定')}</p>
    </div>'''
        
        # 处理数据内容部分
        large_data = len(df) > 1000
        if large_data:
            df_to_display = df.head(1000)
            large_data_notice = "\n    <div class='large-data-notice'>注意：数据集过大，只显示前1000行</div>"
        else:
            df_to_display = df
            large_data_notice = ""
        
        data_table = df_to_display.to_html(index=False)
        
        html_end = f'''\n    <h2>数据内容</h2>{large_data_notice}
    <div class="data-table-container">
        {data_table}
    </div>
</body>
</html>'''
        
        # 拼接所有部分并返回
        return html_start + charts_section + html_end

# 自动化报表引擎优化
class AutoReportEngine:
    """自动化报表引擎（优化版）"""
    
    def _get_data_source(self) -> DataSource:
        """获取数据源实例"""
        data_source_type = self.config.data_source_type.lower()
        
        if data_source_type == 'excel':
            return ExcelDataSource()
        elif data_source_type == 'csv':
            return CSVDataSource()
        elif data_source_type == 'sql':
            return SQLDataSource()
        elif data_source_type == 'api':
            return APIDataSource()
        else:
            raise ValueError(f"不支持的数据源类型: {data_source_type}")
    
    def _get_report_generators(self) -> Dict[str, ReportGenerator]:
        """获取报表生成器实例字典"""
        return {
            'excel': ExcelReportGenerator(),
            'pdf': PDFReportGenerator(),
            'html': HTMLReportGenerator(template_type=self.config.template_type)
        }
    
    def _send_email(self, generated_files: Dict[str, str]):
        """发送邮件"""
        try:
            logger.info("开始发送邮件")
            
            # 检查必要的邮件配置
            smtp_server = config_manager.get('email.smtp_server')
            smtp_port = config_manager.get('email.smtp_port')
            username = config_manager.get('email.username')
            password = config_manager.get('email.password')
            
            if not all([smtp_server, smtp_port, username, password]):
                logger.error("邮件配置不完整，无法发送邮件")
                return
            
            # 创建邮件
            msg = MIMEMultipart()
            msg['From'] = username
            msg['To'] = ', '.join(self.config.recipients)
            msg['Subject'] = f"自动化报表: {self.config.report_name}"
            
            # 添加邮件正文
            body = f"尊敬的用户：\n\n您的报表 '{self.config.report_name}' 已生成完成。\n\n请查看附件中的报表文件。\n\n此邮件由 AutoReport Pro 自动发送。"
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # 添加附件
            for fmt, file_path in generated_files.items():
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
                        msg.attach(part)
            
            # 发送邮件
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                if config_manager.get('email.use_tls', True):
                    server.starttls()
                server.login(username, password)
                server.send_message(msg)
            
            logger.info(f"邮件发送成功，收件人: {', '.join(self.config.recipients)}")
        except Exception as e:
            logger.error(f"发送邮件失败: {e}")
    
    def __init__(self, config: 'ReportConfig'):
        """初始化报表引擎"""
        self.config = config
        
        # 确保输出目录存在
        self.output_dir = config_manager.get('output_dir', 'reports')
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _get_data_source_instance(self, data_source_type: str) -> DataSource:
        """获取数据源实例"""
        data_source_type = data_source_type.lower()
        
        if data_source_type == 'excel':
            return ExcelDataSource()
        elif data_source_type == 'csv':
            return CSVDataSource()
        elif data_source_type == 'sql':
            return SQLDataSource()
        elif data_source_type == 'api':
            return APIDataSource()
        else:
            raise ValueError(f"不支持的数据源类型: {data_source_type}")
    
    def run(self) -> Dict[str, str]:
        """运行报表生成流程（优化版）"""
        try:
            logger.info(f"开始生成报表: {self.config.report_name}")
            
            # 1. 加载数据
            data_frames = {}  # 存储所有数据源的数据
            
            # 兼容旧版本的单数据源配置
            data_sources_to_process = []
            if self.config.data_sources:
                data_sources_to_process = self.config.data_sources
            elif self.config.data_source_type and self.config.data_source_path:
                # 从旧版本字段创建数据源配置
                legacy_source = DataSourceConfig(
                    name="legacy_source",
                    type=self.config.data_source_type,
                    path=self.config.data_source_path,
                    parameters=self.config.parameters or {}
                )
                data_sources_to_process.append(legacy_source)
            
            for ds_config in data_sources_to_process:
                logger.info(f"加载数据源: {ds_config.name or ds_config.type}")
                data_source = self._get_data_source_instance(ds_config.type)
                
                # 创建临时ReportConfig用于加载单个数据源
                temp_config = ReportConfig(
                    report_name=self.config.report_name,
                    output_format=self.config.output_format,
                    schedule=self.config.schedule,
                    recipients=self.config.recipients,
                    template_path=self.config.template_path,
                    filters=self.config.filters,
                    calculations=self.config.calculations,
                    charts=self.config.charts,
                    parameters=ds_config.parameters,  # 使用数据源的参数
                    data_source_type=ds_config.type,  # 兼容旧版
                    data_source_path=ds_config.path  # 兼容旧版
                )
                
                # 加载数据
                df = data_source.load_data(temp_config)
                
                # 验证数据
                if not data_source.validate_data(df):
                    raise ValueError(f"数据源 {ds_config.name or ds_config.type} 验证失败")
                
                data_frames[ds_config.name or f"source_{len(data_frames)}"] = df
            
            # 合并数据
            if len(data_frames) == 0:
                raise ValueError("没有可用的数据源")
            elif len(data_frames) == 1:
                # 只有一个数据源，直接使用
                df = next(iter(data_frames.values()))
            else:
                # 多个数据源，需要合并
                logger.info(f"合并 {len(data_frames)} 个数据源")
                # 这里使用简单的合并策略，实际应用中可能需要更复杂的逻辑
                df = DataProcessor.merge_dataframes(list(data_frames.values()))
            
            # 3. 处理数据
            # 应用筛选
            if self.config.filters:
                df = DataProcessor.apply_filters(df, self.config.filters)
            
            # 应用计算字段
            if self.config.calculations:
                df = DataProcessor.apply_calculations(df, self.config.calculations)
            
            # 4. 计算指标
            metrics = DataProcessor.calculate_metrics(df)
            
            # 5. 生成报表
            generators = self._get_report_generators()
            generated_files = {}
            
            # 对于大型数据集，优先生成轻量级格式
            output_formats = self.config.output_format.copy()
            if len(df) > 100000:
                # 先生成Excel和HTML，再生成PDF（如果需要）
                output_formats = sorted(output_formats, key=lambda x: 0 if x in ['excel', 'html'] else 1)
            
            for fmt in output_formats:
                if fmt == 'email':
                    continue  # 邮件单独处理
                
                generator = generators.get(fmt)
                if generator:
                    # 生成文件名
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    # 根据格式选择正确的文件扩展名
                    if fmt == 'excel':
                        ext = 'xlsx'
                    else:
                        ext = fmt
                    filename = f"{self.config.report_name}_{timestamp}.{ext}"
                    output_path = os.path.join(self.output_dir, filename)
                    
                    # 生成报表
                    generated_path = generator.generate(df, metrics, output_path, self.config.charts)
                    generated_files[fmt] = generated_path
            
            # 6. 发送邮件（如果配置了）
            if 'email' in self.config.output_format and self.config.recipients:
                self._send_email(generated_files)
            
            logger.info(f"报表生成完成: {self.config.report_name}")
            return generated_files
        except Exception as e:
            logger.error(f"生成报表失败: {e}")
            raise

def example_usage():
    """示例用法"""
    print("=== 单数据源示例 ===")
    # 创建单数据源报表配置
    config = ReportConfig(
        report_name="销售报表",
        output_format=["excel", "pdf", "html"],
        data_source_type="excel",
        data_source_path="data/sales_data.xlsx",
        recipients=["user@example.com"],
        filters={
            "销售金额": {"operator": "gt", "value": 1000}
        },
        calculations=[
            {
                "column": "利润率",
                "formula": "(df['销售金额'] - df['成本']) / df['销售金额']"
            }
        ],
        charts=[
            {
                "type": "bar",
                "title": "月度销售趋势",
                "x_field": "月份",
                "y_field": "销售金额"
            }
        ],
        parameters={
            "sheet_name": "Sheet1"
        }
    )
    
    try:
        # 运行报表生成
        engine = AutoReportEngine(config)
        generated_files = engine.run()
        
        print("报表生成成功！")
        for fmt, path in generated_files.items():
            print(f"{fmt}: {path}")
    except FileNotFoundError as e:
        print(f"示例运行提示: {e}")
        print("注意: 要实际运行此示例，需要创建 'data/sales_data.xlsx' 文件")
    except Exception as e:
        print(f"示例运行出错: {e}")
        print("注意: 请确保所有依赖已安装且配置正确")
    
    print("\n=== 多数据源整合示例 ===")
    # 创建多数据源报表配置
    multi_source_config = ReportConfig(
        report_name="多数据源分析报表",
        output_format=["excel", "html"],
        recipients=["analyst@example.com"],
        filters={
            "销售额": {"min": 5000}
        },
        calculations=[
            {
                "column": "销售利润",
                "formula": "df['销售额'] - df['成本']"
            }
        ],
        charts=[
            {
                "type": "line",
                "title": "销售趋势对比",
                "x_field": "日期",
                "y_field": ["销售额", "成本", "销售利润"]
            }
        ],
        # 配置多个数据源
        data_sources=[
            DataSourceConfig(
                name="销售数据",
                type="excel",
                path="data/sales_data.xlsx",
                parameters={"sheet_name": "Sheet1"}
            ),
            DataSourceConfig(
                name="成本数据",
                type="csv",
                path="data/cost_data.csv",
                parameters={"delimiter": ","}
            ),
            DataSourceConfig(
                name="客户数据",
                type="sql",
                path="sqlite:///data/customer.db",
                parameters={
                    "query": "SELECT customer_id, customer_name, region FROM customers"
                }
            )
        ]
    )
    
    try:
        # 运行多数据源报表生成
        multi_source_engine = AutoReportEngine(multi_source_config)
        generated_files = multi_source_engine.run()
        
        print("多数据源报表生成成功！")
        for fmt, path in generated_files.items():
            print(f"{fmt}: {path}")
    except FileNotFoundError as e:
        print(f"多数据源示例运行提示: {e}")
        print("注意: 要实际运行此示例，需要创建相关的数据文件")
    except Exception as e:
        print(f"多数据源示例运行出错: {e}")
        print("注意: 请确保所有依赖已安装且配置正确")

def main():
    """主函数"""
    # 导入更新管理器
    try:
        from update_manager import UpdateManager
        
        # 初始化更新管理器
        update_manager = UpdateManager(
            app_name="AutoReport",
            current_version="1.0.0",  # 这里应该从配置或常量中获取
            update_server_url="https://example.com/updates"  # 这里应该替换为实际的更新服务器URL
        )
        
        # 检查更新
        update_info = update_manager.check_for_updates()
        
        if update_info:
            print(f"发现新版本: {update_info.get('version')}")
            print(f"更新内容: {update_info.get('changelog', '无')}")
            
            # 询问用户是否更新
            user_input = input("是否下载并安装更新？(y/n): ").lower().strip()
            
            if user_input == 'y':
                print("开始下载更新...")
                update_file = update_manager.download_update()
                
                if update_file:
                    print("下载完成，开始安装更新...")
                    if update_manager.install_update(update_file):
                        print("更新安装完成，正在重启应用...")
                        update_manager.restart_application()
                    else:
                        print("更新安装失败，请重试")
                else:
                    print("更新下载失败")
    except ImportError:
        logger.warning("更新管理器模块未找到，跳过更新检查")
    except Exception as e:
        logger.error(f"更新检查失败: {e}")
    
    import argparse
    
    parser = argparse.ArgumentParser(description="自动化报表生成工具")
    parser.add_argument("-c", "--config", type=str, help="配置文件路径")
    parser.add_argument("--name", type=str, help="报表名称")
    parser.add_argument("--source-type", type=str, choices=["excel", "csv", "sql", "api"], help="数据源类型")
    parser.add_argument("--source-path", type=str, help="数据源路径/URL")
    parser.add_argument("--output-format", type=str, nargs="+", choices=["excel", "pdf", "html", "email", "csv"], help="输出格式")
    parser.add_argument("--schedule", type=str, help="调度表达式(cron)")
    parser.add_argument("--recipients", type=str, nargs="+", help="邮件接收者")
    parser.add_argument("--template", type=str, help="模板文件路径")
    parser.add_argument("--example", action="store_true", help="运行示例用法")
    
    args = parser.parse_args()
    
    # 运行示例用法
    if args.example:
        example_usage()
        return
    
    # 从配置文件加载配置
    if args.config:
        # 这里应该实现从配置文件加载ReportConfig的逻辑
        logger.info(f"从配置文件加载: {args.config}")
        # 示例：
        # config = load_config_from_file(args.config)
        # engine = AutoReportEngine(config)
        # engine.run()
    else:
        # 从命令行参数创建配置
        if not all([args.name, args.source_type, args.source_path, args.output_format]):
            parser.print_help()
            return
        
        config = ReportConfig(
            report_name=args.name,
            data_source_type=args.source_type,
            data_source_path=args.source_path,
            output_format=args.output_format,
            schedule=args.schedule,
            recipients=args.recipients or [],
            template_path=args.template
        )
        
        engine = AutoReportEngine(config)
        engine.run()

if __name__ == "__main__":
    main()

